import pandas as pd
import openpyxl
import re
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.drawing.text import (RichTextProperties, ListStyle, Paragraph,
                                    ParagraphProperties, CharacterProperties, RegularTextRun)
from openpyxl.chart.text import RichText as ChartRichText, Text
from openpyxl.chart.title import Title
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import Marker, DataPoint
from openpyxl.drawing.line import LineProperties
import os

MONTHS = {1:'JANUARY',2:'FEBRUARY',3:'MARCH',4:'APRIL',5:'MAY',6:'JUNE',
          7:'JULY',8:'AUGUST',9:'SEPTEMBER',10:'OCTOBER',11:'NOVEMBER',12:'DECEMBER'}

def get_update_date():
    """Generate 'update DD Mon YY' string from today's date"""
    return datetime.now().strftime('update %d %b %y')

TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), 'report_templates')

# ────────────────────────────────────────────────────────────
# TEMPLATE REGISTRY
# เพิ่ม template ใหม่ได้โดยวางไฟล์ .xlsx ใน report_templates/
# แล้วเพิ่ม entry ที่นี่ในรูปแบบ:
#   'key': {'name': 'ชื่อที่แสดงใน UI', 'file': 'ชื่อไฟล์.xlsx'}
# ────────────────────────────────────────────────────────────
TEMPLATES = {
    'room_and_extra': {
        'name': '🛏️ Room and Extra Bed',
        'file': 'template_Room and Extra Bed.xlsx',
        'description': 'รายงานที่มีข้อมูลห้องพักและ Extra Bed (เตียงเสริม) ครบถ้วน'
    },
    'room_only': {
        'name': '🏨 Room Only',
        'file': 'template_Room Only.xlsx',
        'description': 'รายงานเฉพาะห้องพัก ไม่รวม Extra Bed'
    },
    # ── เพิ่ม template ใหม่ด้านล่างนี้ ──
    # 'template3': {
    #     'name': '📋 ชื่อ Template ใหม่',
    #     'file': 'template_ชื่อไฟล์.xlsx',
    #     'description': 'คำอธิบาย template'
    # },
}

DEFAULT_TEMPLATE = 'room_and_extra'


def classify(d):
    d_up = str(d).upper().strip().rstrip('.')
    
    # 1. Check for child-related costs (including standalone 'sharing bed' which implies child)
    if re.search(r'\b(CHD|CHILD|CHILDREN|KID|KIDS)\b', d_up) or re.search(r'\b(SHARE BED|SHARED BED|SHARING BED)\b', d_up):
        if re.search(r'\b(SHARE|SHARED|SHARING|SOFA|UNDER|NO BED)\b', d_up):
            return 'chd_shared'
        else:
            return 'chd_extra'
            
    # 2. Check for adult extra beds
    if re.search(r'\b(EXTRA BED|EX BED|EXT BED|EXT\+BED|EXTRA PERSON|EXTRA PAX|ADT SOFA|SOFA BED)\b', d_up):
        return 'adult_extra'
        
    # 3. Check for general other fees
    other_keywords = [
        'TRANSFER', 'TRANFER', 'BABY COT', 'ADDTIONAL', 'ADDITIONAL',
        'LATE CHECK OUT', 'GALA DINNER', 'SURCHARGE', 'NEW YEAR',
        'MATTRESS', 'VAN', 'SPEED BOAT', 'SPEEDBOAT', 'LONG TAIL BOAT',
        'HALF BOARD', 'FULL BOARD', 'CREDIT', 'MEAL', 'AIRPORT',
        'UPGRADE', 'MASSAGE', 'CHARGE', 'GUARANTEE', 'DINNER', 'FOOD',
        'COMPUL', 'COMPULSORY', 'GALA', 'BOAT', 'CONECING', 'TRAVELLIRI',
        'BENEFIT', 'LATE CHECKOUT', 'BALLOON', 'DECORATION', 'EARLY',
        'WINE', 'SPARKLING', 'SPARKING', 'DRINK', 'BEVERAGE',
        'WELCOME', 'AMENITY', 'FRUIT', 'BASKET', 'CAKE', 'FLOWER'
    ]
    if any(x in d_up for x in other_keywords): return 'other_fee'
    
    # 4. Default to room
    return 'room'


def normalize_name(d, typ):
    if typ == 'chd_extra':   return 'CHD Extra Bed + ABF'
    if typ == 'chd_shared':  return 'CHD Shared Bed + ABF'
    if typ == 'adult_extra': return 'Extra Bed Adult + ABF'
    d_clean = str(d).strip().rstrip('.')
    # Normalize common typos/variants
    d_up = d_clean.upper()
    if 'GRADEN VIEW' in d_up: d_clean = d_clean.replace('Graden','Garden').replace('GRADEN','Garden')
    if d_up == 'ONE BEDROOM REGENCY SUITE + ABF': d_clean = 'One Bedroom Regency Suite + ABF'
    if 'DELUXE POOL VIEW' in d_up: return 'Deluxe Pool View + ABF'
    if 'DELUXE POOL ACCESS' in d_up: return 'Deluxe Pool Access + ABF'
    if 'DELUXE SEA VIEW' in d_up or ('SEA VIEW' in d_up and 'DELUXE' in d_up): return 'Deluxe Sea View + ABF'
    if 'DEKUXE GARDEN' in d_up or 'DELUXE GARDEN' in d_up: return 'Deluxe Garden + ABF'
    if 'SUPERIOR WITH BALCONY' in d_up: return 'Superior with Balcony + ABF'
    if 'SUPERIOR' in d_up and 'BALCONY' not in d_up and 'BANANA' not in d_up: return 'Superior + ABF'
    if 'GRAND DELUXE' in d_up: return 'Grand Deluxe + ABF'
    if 'PREMIUM 2 BEDROOM' in d_up: return 'Premium 2 Bedroom with Roof Deck and Seaview + ABF'
    if 'PREMIUM ROOM' in d_up: return 'Premium Room + ABF'
    if 'STANDARD ROOM' in d_up: return 'Standard Room + ABF'
    if 'EXECUTIVE SUITE' in d_up: return 'Executive Suite + ABF'
    if 'DELUXE MOUNTAIN VIEW' in d_up: return 'Deluxe Mountain View + ABF'
    if 'JUNIOR SUITE MOUNTAIN' in d_up or 'JUNIOR SUITE' in d_up: return 'Junior Suite Mountain View + ABF'
    if 'SKY SUITE' in d_up: return 'Sky Suite Panoramic Ocean View + ABF'
    return d_clean


def thin():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)

def thin_no_bottom():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t)

def make_chart_title(year, total_nights):
    title_str = f'{int(year)} - {total_nights} Nights'
    rpr  = CharacterProperties(sz=1800, b=True, strike='noStrike', kern=1200, baseline=0)
    run  = RegularTextRun(t=title_str); run.rPr = rpr
    pPr  = ParagraphProperties(); pPr.defRPr = rpr
    para = Paragraph(r=[run]); para.pPr = pPr
    bodyPr = RichTextProperties(rot=0, spcFirstLastPara=True, vertOverflow='ellipsis',
                                 vert='horz', wrap='square', anchor='ctr', anchorCtr=True,
                                 noAutofit=False, normAutofit=False, spAutoFit=False)
    rt = ChartRichText(bodyPr=bodyPr, lstStyle=ListStyle(), p=[para])
    return Title(tx=Text(rich=rt), overlay=False)


def generate_report(raw_path: str, hotel_name: str, output_path: str, template_name: str = DEFAULT_TEMPLATE, destination_filter: str = ""):
    """Main function: raw Excel → formatted report Excel
    
    Args:
        raw_path: path to raw input Excel
        hotel_name: hotel name for report header
        output_path: where to save the output Excel
        template_name: key from TEMPLATES dict (default: 'standard')
    """
    # Resolve template path
    if template_name not in TEMPLATES:
        template_name = DEFAULT_TEMPLATE
    template_file = TEMPLATES[template_name]['file']
    template_path = os.path.join(TEMPLATES_DIR, template_file)
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"ไม่พบไฟล์ template: {template_file} (ตรวจสอบโฟลเดอร์ report_templates/)")

    # Load raw data
    df = pd.read_excel(raw_path, header=3)
    
    expected_cols = ['VN','Status','Hotel','Kind','Dest','Country','Agent','AgentID',
                  'Customer','CID','CType','TADT','TCHD','BookingDate','CheckIn',
                  'CheckOut','Details','Nights','Quantity','TotalNights','Cost','TotalCost']
    
    if len(df.columns) != len(expected_cols):
        raise ValueError(f"ไฟล์ Excel ไม่ถูกต้อง! ระบบต้องการ {len(expected_cols)} คอลัมน์ แต่ไฟล์ที่คุณอัปโหลดมี {len(df.columns)} คอลัมน์ (กรุณาเช็กว่าดาวน์โหลด Report ผิดประเภท หรือลืมเลือกบางคอลัมน์ตอน Export หรือไม่)")
        
    df.columns = expected_cols
    df = df[df['Kind'] == 'HOTEL']
    df['CheckIn'] = pd.to_datetime(df['CheckIn'], errors='coerce')
    df['Year']  = df['CheckIn'].dt.year.astype('Int64')
    df['Month'] = df['CheckIn'].dt.month.astype('Int64')

    if destination_filter:
        df = df[df['Dest'].astype(str).str.contains(destination_filter, case=False, na=False)]

    df['Type']     = df['Details'].apply(classify)
    df['RoomName'] = df.apply(lambda r: normalize_name(r['Details'], r['Type']), axis=1)
    years = sorted(df['Year'].dropna().unique())

    # Build monthly_data
    monthly_data = {}
    for year in years:
        ydf = df[df['Year']==year]
        room_df   = ydf[ydf['Type']=='room']
        shared_df = ydf[ydf['Type']=='chd_shared']
        extra_df  = ydf[ydf['Type'].isin(['chd_extra','adult_extra'])]
        monthly_data[year] = {}
        for m in sorted(room_df['Month'].dropna().unique()):
            r = int(room_df[room_df['Month']==m]['TotalNights'].sum())
            s = int(shared_df[shared_df['Month']==m]['TotalNights'].sum())
            e = int(extra_df[extra_df['Month']==m]['TotalNights'].sum())
            monthly_data[year][m] = (r, s, e)

    # Build rooms_data
    rooms_data = {}
    for year in years:
        ydf = df[df['Year']==year]
        if destination_filter:
            summary = ydf.groupby(['Type','Hotel'])['TotalNights'].sum()
        else:
            summary = ydf.groupby(['Type','RoomName'])['TotalNights'].sum()
        rows = []
        
        # Filter types based on template
        allowed_types = ['room','adult_extra','chd_shared','chd_extra','other_fee']
        if template_name == 'room_only':
            allowed_types = ['room']
            
        for t in allowed_types:
            if t in summary.index.get_level_values(0):
                for name, nights in summary[t].sort_values(ascending=False).items():
                    rows.append((t, name, int(nights)))
        rooms_data[year] = rows

    # Build cost_data
    cost_data = {}
    for year in years:
        ydf = df[df['Year']==year]
        room_df = ydf[ydf['Type']=='room']
        other_no_cancel = ydf[
            (ydf['Type']!='room') &
            ~ydf['Status'].astype(str).str.upper().str.contains('CANCEL', na=False)
        ]
        vc       = len(room_df)
        nights   = int(room_df['TotalNights'].sum())
        base_nights = int(room_df['Nights'].sum())
        rc       = room_df['TotalCost'].sum()
        oc       = other_no_cancel['TotalCost'].sum()
        avg      = round(base_nights/vc, 2) if vc else 0
        adt      = int(room_df['TADT'].sum())
        chd      = int(room_df['TCHD'].sum())
        adt_only = int((room_df['TCHD']==0).sum())
        adt_chd  = int((room_df['TCHD']>0).sum())
        cost_data[year] = (vc, nights, rc, oc, avg, adt, chd, adt_only, adt_chd)

    # ===== BUILD WORKBOOK =====
    ROOM_COL  = 15
    CHART_COL = 9

    gold_fill   = PatternFill('solid', fgColor='FFC000')
    purple_fill = PatternFill('solid', fgColor='BE87FB')

    f_title     = Font(name='Calibri', bold=True, size=16)
    f_update    = Font(name='Calibri', bold=True, size=11)
    f_year_hdr  = Font(name='Calibri', bold=True, size=11)
    f_col_hdr   = Font(name='Calibri', bold=False, size=11)
    f_data      = Font(name='Calibri', bold=False, size=11)
    f_cost_hdr  = Font(name='Calibri', bold=True, size=16)
    f_year_cost = Font(name='Calibri', bold=True, size=16)
    f_data_cost = Font(name='Calibri', bold=False, size=14)
    al_cc = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al_c  = Alignment(horizontal='center')

    tmpl = openpyxl.load_workbook(template_path)
    wb = openpyxl.Workbook()
    ws_report = wb.active; ws_report.title = 'report'
    ws_cost = wb.create_sheet('cost')

    for col, dim in tmpl.worksheets[0].column_dimensions.items():
        ws_report.column_dimensions[col].width = dim.width
    if len(tmpl.worksheets) > 1:
        for col, dim in tmpl.worksheets[1].column_dimensions.items():
            ws_cost.column_dimensions[col].width = dim.width

    # Enforce width for Room Name column and wrap text
    ws_report.column_dimensions[openpyxl.utils.cell.get_column_letter(ROOM_COL)].width = 50
    ws_report.column_dimensions['M'].width = 8.5
    ws_report.column_dimensions['S'].width = 8.5
    ws_report.column_dimensions['W'].width = 8.5
    
    ws_cost_widths = {'D':8, 'E':12, 'F':14, 'G':16, 'H':20, 'I':20, 'J':14, 'K':8, 'L':8, 'M':20, 'N':8, 'O':20, 'P':8}
    for c_letter, w in ws_cost_widths.items():
        ws_cost.column_dimensions[c_letter].width = w

    # ----- REPORT SHEET -----
    c = ws_report['K3']
    c.value = hotel_name
    c.font = Font(name='Calibri', bold=True, size=36)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    try:
        ws_report.merge_cells('K3:T3')
    except Exception:
        pass

    c = ws_report['T4']
    c.value = get_update_date()
    c.font = f_update

    left_row = 7
    chart_top_row = 5  # 0-indexed

    for year in years:
        months = monthly_data[year]
        has_shared = any(v[1] for v in months.values())
        has_extra  = any(v[2] for v in months.values())
        
        if template_name == 'room_only':
            has_shared = False
            has_extra = False
        header_row = left_row

        yc = ws_report.cell(left_row, 1)
        yc.value = int(year); yc.font = f_year_hdr
        if year == 2026: yc.fill = purple_fill
        ws_report.cell(left_row,2).value = 'total nights'; ws_report.cell(left_row,2).font = f_col_hdr
        if has_shared:
            ws_report.cell(left_row,3).value = 'Shared Bed'; ws_report.cell(left_row,3).font = f_col_hdr
        if has_extra:
            ws_report.cell(left_row,4).value = 'Extra Bed';  ws_report.cell(left_row,4).font = f_col_hdr
        left_row += 1
        data_start = left_row

        for m in sorted(months.keys()):
            r, s, e = months[m]
            ws_report.cell(left_row,1).value = MONTHS[m]; ws_report.cell(left_row,1).font = f_data
            ws_report.cell(left_row,2).value = r;         ws_report.cell(left_row,2).font = f_data
            if has_shared:
                ws_report.cell(left_row,3).value = s if s else 0
                ws_report.cell(left_row,3).font = f_data
            if has_extra:
                ws_report.cell(left_row,4).value = e if e else 0
                ws_report.cell(left_row,4).font = f_data
            left_row += 1

        data_end = left_row - 1
        left_row += 2

        # Chart
        chart_bottom_row = chart_top_row + 24
        chart = LineChart()
        chart.title = make_chart_title(year, sum(v[0] for v in months.values()))
        chart.style = 10; chart.width = 15; chart.height = 7.5
        chart.add_data(Reference(ws_report,min_col=2,min_row=header_row,max_row=data_end),
                       titles_from_data=True)
        if has_shared:
            chart.add_data(Reference(ws_report,min_col=3,min_row=header_row,max_row=data_end),
                           titles_from_data=True)
        if has_extra:
            chart.add_data(Reference(ws_report,min_col=4,min_row=header_row,max_row=data_end),
                           titles_from_data=True)
        chart.set_categories(Reference(ws_report,min_col=1,min_row=data_start,max_row=data_end))
        
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.y_axis.tickLblPos = "none"
        chart.y_axis.spPr = GraphicalProperties(ln=LineProperties(noFill=True))
        chart.roundedCorners = True
        chart.graphicalProperties = GraphicalProperties(solidFill='E7E6E6')
        
        if template_name == 'room_and_extra':
            if chart.legend:
                chart.legend.position = 'r'
        else:
            chart.legend = None # Remove legend to match reference image
        
        # Consistent colors to match reference image: 0:Total(Green), 1:Shared(Blue), 2:Extra(Yellow)
        series_colors = ['70AD47', '5B9BD5', 'FFC000']
        for i, s in enumerate(chart.series):
            s.graphicalProperties.line.solidFill = series_colors[i % len(series_colors)]
            s.graphicalProperties.line.width = 25000 # ~2pt
            s.smooth = False
            s.marker = Marker(symbol='circle', size=15)
            
            if i == 0:
                pt_colors = ['5B9BD5', 'ED7D31', '70AD47', 'FFC000', '44546A', '8FAADC', 'F4B084']
                for pt_idx in range(len(months)):
                    pt = DataPoint(idx=pt_idx)
                    pt.marker = Marker(symbol='circle', size=15)
                    pt.marker.graphicalProperties.solidFill = pt_colors[pt_idx % len(pt_colors)]
                    s.dPt.append(pt)
            
            # Data Labels
            lbl = DataLabelList()
            lbl.showVal        = True
            lbl.showCatName    = False
            lbl.showSerName    = False
            lbl.showLegendKey  = False
            lbl.showPercent    = False
            lbl.showBubbleSize = False
            # Position labels inside marker
            lbl.dLblPos        = 'ctr'
            
            cp = CharacterProperties(solidFill='FFFFFF')
            lbl.txPr = ChartRichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            
            s.dLbls = lbl
        
        anchor = TwoCellAnchor()
        anchor._from = AnchorMarker(col=CHART_COL, colOff=0, row=chart_top_row, rowOff=0)
        anchor.to    = AnchorMarker(col=CHART_COL+13, colOff=29483, row=chart_bottom_row, rowOff=90298)
        chart.anchor = anchor
        ws_report.add_chart(chart)

        # Room name table below chart
        room_excel_row = chart_bottom_row + 4
        c1 = ws_report.cell(room_excel_row, ROOM_COL)
        c1.value = 'Hotel Name' if destination_filter else 'Room name'; c1.font = f_year_hdr; c1.fill = gold_fill; c1.border = thin_no_bottom()
        c2 = ws_report.cell(room_excel_row, ROOM_COL+1)
        c2.value = int(year); c2.font = f_year_hdr; c2.fill = gold_fill
        c2.alignment = al_c; c2.border = thin_no_bottom()
        room_excel_row += 1

        for rtype, name, nights in rooms_data[year]:
            c1 = ws_report.cell(room_excel_row, ROOM_COL)
            c2 = ws_report.cell(room_excel_row, ROOM_COL+1)
            c1.value = name; c1.font = f_data; c1.border = thin(); c1.alignment = Alignment(wrap_text=True)
            c2.value = nights; c2.font = f_data; c2.alignment = al_c; c2.border = thin()
            room_excel_row += 1

        chart_top_row = room_excel_row + 1

    # ----- COST SHEET -----
    c = ws_cost['E7']
    c.value = hotel_name
    c.font = Font(name='Calibri', bold=True, size=36)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_cost.merge_cells('E7:P7')

    c = ws_cost['O8']
    c.value = get_update_date(); c.font = f_update
    c.alignment = Alignment(horizontal='right')

    hdr_cols = {
        'E10':'Total VC', 'F10':'Total Nights', 'G10':'Total Cost',
        'H10':'Total Cost - Nights', 'I10':'Total Cost - Other', 'J10':'Average Nights',
        'K10':'ADT', 'L10':'CHD', 'M10':'Total VC - ADT ONLY', 'N10':'%',
        'O10':'Total VC - ADT+CHD', 'P10':'%'
    }
    for addr, val in hdr_cols.items():
        c = ws_cost[addr]; c.value = val; c.font = f_cost_hdr; c.alignment = al_cc
        m = Side(style='medium')
        c.border = Border(left=m, right=m, top=m, bottom=m)

    t_side = Side(style='thin'); m_side = Side(style='medium')
    for i, year in enumerate(years):
        row = 11 + i
        vc,nights,rc,oc,avg,adt,chd,adt_only,adt_chd = cost_data[year]
        top_s = t_side if i > 0 else m_side
        bot_s = m_side if i == len(years)-1 else t_side

        ws_cost.cell(row,4).value = int(year); ws_cost.cell(row,4).font = f_year_cost
        ws_cost.cell(row,4).border = Border(left=m_side, right=t_side, top=top_s, bottom=bot_s)

        data_vals = [
            vc, nights, f'=H{row}+I{row}', round(rc,2), round(oc,2), avg,
            adt, chd, adt_only, f'=M{row}*100/E{row}', adt_chd, f'=O{row}*100/E{row}'
        ]
        for col, val in zip(range(5,17), data_vals):
            c = ws_cost.cell(row, col); c.value = val; c.font = f_data_cost
            left_b  = m_side if col == 13 else t_side
            right_b = m_side if col == 16 else t_side
            c.border = Border(left=left_b, right=right_b, top=top_s, bottom=bot_s)
            # Number formats
            if col in (7, 8, 9):   # Total Cost, Total Cost-Nights, Total Cost-Other
                c.number_format = '#,##0.00'
            elif col in (14, 16):  # % columns
                c.number_format = '0.00'

    # 3 Line charts in cost sheet (same format as report sheet)
    if template_name != 'room_only':
        data_row_end = 10 + len(years)
        CHART_START_ROW = 17
        for title, data_col, color, chart_col, is_cost in [
            ('Total Nights per Year', 6, '2E75B6', 4,  False),
            ('Total VC per Year',     5, '70AD47', 13, False),
            ('Total Cost per Year',   7, 'ED7D31', 22, True),
        ]:
            chart = LineChart(); chart.title=title; chart.style=10
            chart.width=10; chart.height=8; chart.legend=None
            chart.add_data(Reference(ws_cost,min_col=data_col,min_row=10,max_row=data_row_end),
                           titles_from_data=True)
            chart.set_categories(Reference(ws_cost,min_col=4,min_row=11,max_row=data_row_end))
            chart.series[0].graphicalProperties.line.solidFill = color
            chart.series[0].graphicalProperties.line.width = 25000
            
            chart.x_axis.delete = False
            chart.y_axis.delete = False
            for s in chart.series:
                s.smooth = False
                lbl = DataLabelList()
                lbl.showVal        = True
                lbl.showCatName    = False
                lbl.showSerName    = False
                lbl.showLegendKey  = False
                lbl.showPercent    = False
                lbl.showBubbleSize = False
                if is_cost:
                    lbl.numFmt = '#,##0.00'
                    lbl.sourceLinked = False
                s.dLbls = lbl
            
            anchor = TwoCellAnchor()
            anchor._from = AnchorMarker(col=chart_col, colOff=0, row=CHART_START_ROW, rowOff=0)
            anchor.to    = AnchorMarker(col=chart_col+8, colOff=0, row=CHART_START_ROW+18, rowOff=0)
            chart.anchor = anchor
            ws_cost.add_chart(chart)

    wb.save(output_path)
    return output_path
